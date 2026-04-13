#!/usr/bin/env python3
"""Build a product-image alignment workbook from local image assets and json/jsonl data."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import os
import re
import sys
import traceback
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import urlparse

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover - import guard
    raise SystemExit(
        "openpyxl is required to export xlsx files. "
        "Please install it in the current Python environment."
    ) from exc

try:
    from PIL import Image as PILImage
    from PIL import Image
except Exception as exc:  # pragma: no cover - import guard
    raise SystemExit(
        "Pillow is required to embed image previews into the xlsx workbook."
    ) from exc

warnings.filterwarnings("ignore", category=Image.DecompressionBombWarning)


IMAGE_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
    ".bmp",
    ".gif",
    ".tif",
    ".tiff",
    ".heic",
    ".heif",
}

IGNORE_DIR_NAMES = {
    ".git",
    ".omx",
    ".idea",
    ".vscode",
    "__pycache__",
    "node_modules",
    ".DS_Store",
}

TYPE_KEYWORDS = {
    "cover": ["主图", "封面", "首图", "cover", "thumb", "头图"],
    "gallery": ["轮播", "组图", "banner", "gallery", "通用", "images", "图集", "展示图"],
    "detail": ["详情", "详情页", "detail", "长图"],
    "sku": ["sku", "规格", "款式", "色号", "容量", "型号", "spec"],
    "poster": ["海报", "poster", "宣传", "画报", "推广图"],
}

GENERIC_ALIAS_STOPWORDS = {
    "主图",
    "详情",
    "详情页",
    "海报",
    "封面",
    "组图",
    "通用",
    "images",
    "image",
    "banner",
    "poster",
    "cover",
    "detail",
    "sku",
    "spec",
    "产品",
    "商品",
}

PRODUCT_FIELD_CANDIDATES = {
    "id": ["id", "_legacy_id", "_id", "product_id", "spu_id"],
    "name": ["name", "product_name", "title", "spu_name"],
    "description": ["description", "desc", "subtitle", "summary"],
    "category_id": ["category_id", "categoryId", "category", "cat_id"],
    "price": ["retail_price", "min_price", "sale_price", "price", "member_price"],
    "original_price": ["market_price", "original_price", "line_price"],
    "stock": ["stock", "inventory", "qty", "quantity"],
    "images": ["images", "gallery_images", "slider_images"],
    "detail_images": ["detail_images", "details", "detail", "long_images"],
    "cover_image": ["image", "cover_image", "cover", "main_image", "thumbnail", "thumb"],
}

SKU_FIELD_CANDIDATES = {
    "sku_id": ["id", "_legacy_id", "_id", "sku_id"],
    "product_id": ["product_id", "spu_id"],
    "name": ["name", "title", "spec", "spec_value"],
    "price": ["price", "retail_price", "sale_price"],
    "original_price": ["original_price", "market_price"],
    "stock": ["stock", "inventory", "qty"],
    "image": ["image", "cover_image", "thumb", "thumbnail"],
}


@dataclass
class ImageAsset:
    path: str
    relative_path: str
    file_name: str
    stem: str
    ext: str
    size_bytes: int
    sha1: str
    detected_type: str
    search_text: str
    parent_dir: str
    raw_text: str
    width: Optional[int] = None
    height: Optional[int] = None


@dataclass
class SkuRecord:
    sku_id: str
    product_id: str
    name: str = ""
    price: Optional[float] = None
    original_price: Optional[float] = None
    stock: Optional[float] = None
    image_urls: List[str] = field(default_factory=list)
    source_json_file: str = ""
    source_priority: int = 0


@dataclass
class ProductRecord:
    product_id: str
    product_name: str = ""
    description: str = ""
    price: Optional[float] = None
    original_price: Optional[float] = None
    stock: Optional[float] = None
    category_id: str = ""
    category_name: str = ""
    existing_cover_images: List[str] = field(default_factory=list)
    existing_gallery_images: List[str] = field(default_factory=list)
    existing_detail_images: List[str] = field(default_factory=list)
    source_json_file: str = ""
    source_priority: int = 0
    merged_sources: List[str] = field(default_factory=list)
    sku_ids: List[str] = field(default_factory=list)
    sku_image_urls: List[str] = field(default_factory=list)
    aliases: List[str] = field(default_factory=list)
    unique_aliases: List[str] = field(default_factory=list)
    remote_file_stems: List[str] = field(default_factory=list)


@dataclass
class MatchCandidate:
    product_id: str
    score: float
    confidence: str
    basis: List[str]


@dataclass
class ImageAssignment:
    image: ImageAsset
    product_id: Optional[str]
    score: float
    confidence: str
    basis: List[str]
    top_candidates: List[MatchCandidate]


@dataclass
class ImageEmbed:
    row_number: int
    column_name: str
    image_path: str


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def source_priority(path: Path) -> int:
    path_text = str(path).replace("\\", "/")
    if "cloud-mp/mysql/jsonl" in path_text:
        return 300
    if "cloud-mp/cloudbase-import" in path_text:
        return 200
    if "cloud-mp/cloudbase-seed" in path_text:
        return 100
    return 0


def canonical_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(int(value))
    if isinstance(value, (int, float)):
        if int(value) == value:
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text


def normalize_text(value: str) -> str:
    if not value:
        return ""
    text = str(value).strip().lower()
    text = text.replace("\\", "/")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[_\-]+", "", text)
    text = re.sub(r"[^\w\u4e00-\u9fff/]+", "", text)
    return text


def split_text_segments(value: str) -> List[str]:
    if not value:
        return []
    raw_parts = re.split(r"[\/\\|+,_\-\s（）()【】\[\]·、，,：:;；]+", str(value))
    result: List[str] = []
    for part in raw_parts:
        part = part.strip()
        if not part:
            continue
        if part in GENERIC_ALIAS_STOPWORDS:
            continue
        result.append(part)
    return result


def alias_candidates_from_text(value: str) -> List[str]:
    aliases: List[str] = []
    for part in split_text_segments(value):
        normalized = normalize_text(part)
        if len(normalized) < 2:
            continue
        if normalized.isdigit():
            continue
        aliases.append(normalized)
        if re.search(r"[\u4e00-\u9fff]", normalized) and len(normalized) >= 4:
            for length in range(2, min(6, len(normalized)) + 1):
                for start in range(0, len(normalized) - length + 1):
                    sub = normalized[start : start + length]
                    if sub in GENERIC_ALIAS_STOPWORDS or len(sub) < 2 or sub.isdigit():
                        continue
                    aliases.append(sub)
    deduped: List[str] = []
    seen = set()
    for alias in aliases:
        if alias in seen:
            continue
        seen.add(alias)
        deduped.append(alias)
    return deduped


def is_url(value: str) -> bool:
    return isinstance(value, str) and value.startswith(("http://", "https://"))


def collect_string_list(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        result: List[str] = []
        for item in value:
            result.extend(collect_string_list(item))
        return result
    if isinstance(value, dict):
        return []
    if isinstance(value, (int, float, bool)):
        return [str(value)]
    text = str(value).strip()
    if not text:
        return []
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return [str(item).strip() for item in parsed if str(item).strip()]
        except Exception:
            pass
    return [text]


def pick_first(record: Dict[str, Any], keys: Sequence[str]) -> Any:
    for key in keys:
        if key in record and record[key] not in (None, "", []):
            return record[key]
    return None


def unique_preserve(values: Iterable[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for value in values:
        if not value:
            continue
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def parse_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_money(value: Any, field_name: str, source_path: str) -> Optional[float]:
    number = parse_number(value)
    if number is None:
        return None
    if "cloudbase" in source_path.replace("\\", "/") and field_name in {
        "min_price",
        "original_price",
        "price",
    }:
        if abs(number) >= 1000:
            number = number / 100.0
    return round(number, 2)


def extract_url_basename(value: str) -> str:
    if not value:
        return ""
    parsed = urlparse(value)
    path = parsed.path or value
    return os.path.basename(path)


def extract_url_stem(value: str) -> str:
    basename = extract_url_basename(value)
    return Path(basename).stem


def detect_image_type(path: Path) -> str:
    lower = str(path).lower()
    for image_type, keywords in TYPE_KEYWORDS.items():
        if any(keyword.lower() in lower for keyword in keywords):
            return image_type
    return "unknown"


def sha1_file(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_image_dimensions(path: Path) -> Tuple[Optional[int], Optional[int]]:
    try:
        with PILImage.open(path) as img:
            width, height = img.size
            return int(width), int(height)
    except Exception:
        return None, None


def configure_logging(output_dir: Path, verbose: bool) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    log_level = logging.DEBUG if verbose else logging.INFO
    log_file = output_dir / "product_image_mapping.log"
    logging.basicConfig(
        level=log_level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(log_file, encoding="utf-8"),
        ],
    )


def scan_images(image_roots: Sequence[Path]) -> List[ImageAsset]:
    image_assets: List[ImageAsset] = []
    for root in image_roots:
        if not root.exists():
            logging.warning("Image root does not exist: %s", root)
            continue
        logging.info("Scanning images under %s", root)
        for dirpath, dirnames, filenames in os.walk(root):
            dirnames[:] = [
                name
                for name in dirnames
                if name not in IGNORE_DIR_NAMES and not name.startswith(".")
            ]
            for filename in filenames:
                file_path = Path(dirpath) / filename
                ext = file_path.suffix.lower()
                if ext not in IMAGE_EXTENSIONS:
                    continue
                try:
                    relative_path = str(file_path.relative_to(root))
                except Exception:
                    relative_path = str(file_path)
                try:
                    width, height = read_image_dimensions(file_path)
                    image_assets.append(
                        ImageAsset(
                            path=str(file_path.resolve()),
                            relative_path=relative_path,
                            file_name=file_path.name,
                            stem=file_path.stem,
                            ext=ext,
                            size_bytes=file_path.stat().st_size,
                            sha1=sha1_file(file_path),
                            detected_type=detect_image_type(file_path),
                            search_text=normalize_text(
                                f"{relative_path}/{file_path.parent.name}/{file_path.stem}"
                            ),
                            parent_dir=file_path.parent.name,
                            raw_text=str(file_path).lower(),
                            width=width,
                            height=height,
                        )
                    )
                except Exception:
                    logging.exception("Failed to scan image file: %s", file_path)
    logging.info("Scanned %s image files", len(image_assets))
    return image_assets


def should_skip_data_dir(path: Path, output_dir: Path) -> bool:
    name = path.name
    if name in IGNORE_DIR_NAMES:
        return True
    if name.startswith(".") and name not in {".config"}:
        return True
    try:
        output_dir.resolve().relative_to(path.resolve())
        return True
    except Exception:
        return False


def load_records(path: Path) -> List[Dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8-sig") as handle:
            try:
                parsed = json.load(handle)
                if isinstance(parsed, list):
                    return [item for item in parsed if isinstance(item, dict)]
                if isinstance(parsed, dict):
                    return [parsed]
                return []
            except json.JSONDecodeError:
                pass
    except Exception:
        logging.exception("Failed to read json file: %s", path)
        return []

    path_text = str(path).replace("\\", "/").lower()
    should_try_jsonl = path.suffix.lower() == ".jsonl" or any(
        marker in path_text for marker in ("/jsonl/", "/cloudbase-import/")
    )
    if not should_try_jsonl:
        return []

    records: List[Dict[str, Any]] = []
    try:
        with path.open("r", encoding="utf-8-sig") as handle:
            for line_number, line in enumerate(handle, start=1):
                line = line.strip()
                if not line:
                    continue
                try:
                    parsed = json.loads(line)
                    if isinstance(parsed, dict):
                        records.append(parsed)
                except json.JSONDecodeError as exc:
                    logging.warning(
                        "Skip malformed jsonl line %s in %s: %s",
                        line_number,
                        path,
                        exc,
                    )
    except Exception:
        logging.exception("Failed to parse jsonl-like file: %s", path)
    return records


def infer_dataset_kind(path: Path, records: Sequence[Dict[str, Any]]) -> str:
    if not records:
        return "unknown"
    stem = path.stem.lower()
    if stem == "categories":
        return "categories"
    if stem in {"skus", "product_skus"}:
        return "skus"
    if stem == "products":
        return "products"

    sample = records[: min(5, len(records))]
    if all(
        any(key in record for key in ("name", "title", "product_name"))
        and any(key in record for key in ("retail_price", "min_price"))
        and any(key in record for key in ("images", "detail_images", "category_id"))
        for record in sample
    ):
        return "products"
    if all(
        "product_id" in record and any(key in record for key in ("price", "retail_price"))
        and any(key in record for key in ("image", "spec", "sku_code", "spec_name"))
        for record in sample
    ):
        return "skus"
    if all(
        any(key in record for key in ("name", "title"))
        and "parent_id" in record
        and any(key in record for key in ("level", "sort_order"))
        and not any(key in record for key in ("retail_price", "price", "min_price"))
        for record in sample
    ):
        return "categories"
    return "unknown"


def scan_data_files(data_roots: Sequence[Path], output_dir: Path) -> Tuple[List[Path], Dict[str, List[Path]]]:
    all_files: List[Path] = []
    datasets: Dict[str, List[Path]] = defaultdict(list)
    for root in data_roots:
        if not root.exists():
            logging.warning("Data root does not exist: %s", root)
            continue
        logging.info("Scanning json/jsonl under %s", root)
        for dirpath, dirnames, filenames in os.walk(root):
            dir_path = Path(dirpath)
            dirnames[:] = [
                name
                for name in dirnames
                if not should_skip_data_dir(dir_path / name, output_dir)
            ]
            for filename in filenames:
                file_path = dir_path / filename
                if file_path.suffix.lower() not in {".json", ".jsonl"}:
                    continue
                all_files.append(file_path)
                records = load_records(file_path)
                kind = infer_dataset_kind(file_path, records)
                if kind != "unknown":
                    datasets[kind].append(file_path)
    for kind, files in datasets.items():
        logging.info("Detected %s dataset files: %s", kind, len(files))
    return all_files, datasets


def merge_category_maps(category_files: Sequence[Path]) -> Dict[str, str]:
    category_map: Dict[str, Tuple[int, str]] = {}
    for path in sorted(category_files, key=source_priority, reverse=True):
        records = load_records(path)
        priority = source_priority(path)
        for record in records:
            category_id = canonical_id(pick_first(record, ["id", "_legacy_id", "_id"]))
            name = str(pick_first(record, ["name", "title"]) or "").strip()
            if not category_id or not name:
                continue
            current = category_map.get(category_id)
            if current is None or priority >= current[0]:
                category_map[category_id] = (priority, name)
    return {key: value for key, (_, value) in category_map.items()}


def normalize_product_record(record: Dict[str, Any], path: Path, category_map: Dict[str, str]) -> Optional[ProductRecord]:
    product_id = canonical_id(pick_first(record, PRODUCT_FIELD_CANDIDATES["id"]))
    product_name = str(pick_first(record, PRODUCT_FIELD_CANDIDATES["name"]) or "").strip()
    if not product_id or not product_name:
        return None

    source_path = str(path.resolve())
    category_id = canonical_id(pick_first(record, PRODUCT_FIELD_CANDIDATES["category_id"]))
    price_value = None
    for key in PRODUCT_FIELD_CANDIDATES["price"]:
        if key in record:
            price_value = normalize_money(record.get(key), key, source_path)
            if price_value is not None:
                break
    original_price_value = None
    for key in PRODUCT_FIELD_CANDIDATES["original_price"]:
        if key in record:
            original_price_value = normalize_money(record.get(key), key, source_path)
            if original_price_value is not None:
                break
    stock_value = None
    for key in PRODUCT_FIELD_CANDIDATES["stock"]:
        if key in record:
            stock_value = parse_number(record.get(key))
            if stock_value is not None:
                break

    cover_images = collect_string_list(pick_first(record, PRODUCT_FIELD_CANDIDATES["cover_image"]))
    gallery_images = collect_string_list(pick_first(record, PRODUCT_FIELD_CANDIDATES["images"]))
    detail_images = collect_string_list(pick_first(record, PRODUCT_FIELD_CANDIDATES["detail_images"]))

    return ProductRecord(
        product_id=product_id,
        product_name=product_name,
        description=str(pick_first(record, PRODUCT_FIELD_CANDIDATES["description"]) or "").strip(),
        price=price_value,
        original_price=original_price_value,
        stock=stock_value,
        category_id=category_id,
        category_name=category_map.get(category_id, ""),
        existing_cover_images=unique_preserve(cover_images),
        existing_gallery_images=unique_preserve(gallery_images),
        existing_detail_images=unique_preserve(detail_images),
        source_json_file=source_path,
        source_priority=source_priority(path),
        merged_sources=[source_path],
    )


def normalize_sku_record(record: Dict[str, Any], path: Path) -> Optional[SkuRecord]:
    sku_id = canonical_id(pick_first(record, SKU_FIELD_CANDIDATES["sku_id"]))
    product_id = canonical_id(pick_first(record, SKU_FIELD_CANDIDATES["product_id"]))
    if not product_id:
        return None
    source_path = str(path.resolve())
    price_value = None
    for key in SKU_FIELD_CANDIDATES["price"]:
        if key in record:
            price_value = normalize_money(record.get(key), key, source_path)
            if price_value is not None:
                break
    original_price_value = None
    for key in SKU_FIELD_CANDIDATES["original_price"]:
        if key in record:
            original_price_value = normalize_money(record.get(key), key, source_path)
            if original_price_value is not None:
                break
    stock_value = None
    for key in SKU_FIELD_CANDIDATES["stock"]:
        if key in record:
            stock_value = parse_number(record.get(key))
            if stock_value is not None:
                break
    image_values: List[str] = []
    for key in SKU_FIELD_CANDIDATES["image"]:
        if key in record:
            image_values.extend(collect_string_list(record.get(key)))
    return SkuRecord(
        sku_id=sku_id or f"sku:{product_id}:{hash(json.dumps(record, ensure_ascii=False, sort_keys=True))}",
        product_id=product_id,
        name=str(pick_first(record, SKU_FIELD_CANDIDATES["name"]) or "").strip(),
        price=price_value,
        original_price=original_price_value,
        stock=stock_value,
        image_urls=unique_preserve(image_values),
        source_json_file=source_path,
        source_priority=source_priority(path),
    )


def merge_products(product_files: Sequence[Path], category_map: Dict[str, str]) -> Dict[str, ProductRecord]:
    merged: Dict[str, ProductRecord] = {}
    for path in sorted(product_files, key=source_priority, reverse=True):
        records = load_records(path)
        for record in records:
            product = normalize_product_record(record, path, category_map)
            if product is None:
                continue
            current = merged.get(product.product_id)
            if current is None:
                merged[product.product_id] = product
                continue
            current.merged_sources = unique_preserve(current.merged_sources + product.merged_sources)
            if not current.product_name and product.product_name:
                current.product_name = product.product_name
            if not current.description and product.description:
                current.description = product.description
            if current.price is None and product.price is not None:
                current.price = product.price
            if current.original_price is None and product.original_price is not None:
                current.original_price = product.original_price
            if current.stock is None and product.stock is not None:
                current.stock = product.stock
            if not current.category_id and product.category_id:
                current.category_id = product.category_id
            if not current.category_name and product.category_name:
                current.category_name = product.category_name
            current.existing_cover_images = unique_preserve(
                current.existing_cover_images + product.existing_cover_images
            )
            current.existing_gallery_images = unique_preserve(
                current.existing_gallery_images + product.existing_gallery_images
            )
            current.existing_detail_images = unique_preserve(
                current.existing_detail_images + product.existing_detail_images
            )
    return merged


def merge_skus(sku_files: Sequence[Path]) -> Dict[str, SkuRecord]:
    merged: Dict[str, SkuRecord] = {}
    for path in sorted(sku_files, key=source_priority, reverse=True):
        records = load_records(path)
        for record in records:
            sku = normalize_sku_record(record, path)
            if sku is None:
                continue
            current = merged.get(sku.sku_id)
            if current is None:
                merged[sku.sku_id] = sku
                continue
            current.image_urls = unique_preserve(current.image_urls + sku.image_urls)
            if current.price is None and sku.price is not None:
                current.price = sku.price
            if current.original_price is None and sku.original_price is not None:
                current.original_price = sku.original_price
            if current.stock is None and sku.stock is not None:
                current.stock = sku.stock
            if not current.name and sku.name:
                current.name = sku.name
    return merged


def attach_skus_to_products(products: Dict[str, ProductRecord], skus: Dict[str, SkuRecord]) -> None:
    for sku in skus.values():
        product = products.get(sku.product_id)
        if product is None:
            continue
        product.sku_ids = unique_preserve(product.sku_ids + [sku.sku_id])
        product.sku_image_urls = unique_preserve(product.sku_image_urls + sku.image_urls)
        if product.price is None and sku.price is not None:
            product.price = sku.price
        if product.original_price is None and sku.original_price is not None:
            product.original_price = sku.original_price
        if product.stock is None and sku.stock is not None:
            product.stock = sku.stock


def build_aliases(products: Dict[str, ProductRecord]) -> None:
    alias_owner_counts: Counter[str] = Counter()
    per_product_aliases: Dict[str, List[str]] = {}

    for product in products.values():
        aliases = alias_candidates_from_text(product.product_name)
        aliases.extend(alias_candidates_from_text(product.description))
        aliases = [
            alias
            for alias in unique_preserve(aliases)
            if alias not in GENERIC_ALIAS_STOPWORDS and len(alias) >= 2
        ]
        per_product_aliases[product.product_id] = aliases
        alias_owner_counts.update(set(aliases))

    for product in products.values():
        aliases = per_product_aliases.get(product.product_id, [])
        unique_aliases = [
            alias
            for alias in aliases
            if alias_owner_counts[alias] == 1 and len(alias) >= 2
        ]
        remote_stems = [
            extract_url_stem(url)
            for url in (
                product.existing_cover_images
                + product.existing_gallery_images
                + product.existing_detail_images
                + product.sku_image_urls
            )
        ]
        product.aliases = sorted(aliases, key=len, reverse=True)
        product.unique_aliases = sorted(unique_aliases, key=len, reverse=True)
        product.remote_file_stems = unique_preserve(
            [normalize_text(stem) for stem in remote_stems if stem]
        )


def score_image_for_product(image: ImageAsset, product: ProductRecord) -> Tuple[float, List[str]]:
    score = 0.0
    reasons: List[str] = []

    image_stem = normalize_text(image.stem)
    image_name = normalize_text(image.file_name)
    if image_stem and image_stem in product.remote_file_stems:
        score += 120
        reasons.append(f"remote_basename:{image_stem}")
    elif image_name and image_name in product.remote_file_stems:
        score += 120
        reasons.append(f"remote_filename:{image_name}")

    for alias in product.unique_aliases:
        if alias and alias in image.search_text:
            alias_score = min(30 + len(alias) * 7, 90)
            if normalize_text(image.parent_dir) == alias:
                alias_score += 10
            score += alias_score
            reasons.append(f"unique_alias:{alias}")
            break

    for alias in product.aliases[:8]:
        if alias and alias in image.search_text and f"unique_alias:{alias}" not in reasons:
            alias_score = min(8 + len(alias) * 4, 35)
            score += alias_score
            reasons.append(f"alias:{alias}")
            break

    if product.product_id:
        pattern = rf"(product|sku|spu|商品|产品|id)[_\-/\s]*{re.escape(product.product_id)}([^0-9]|$)"
        if re.search(pattern, image.raw_text):
            score += 18
            reasons.append(f"product_id:{product.product_id}")

    name_similarity = SequenceMatcher(
        None,
        normalize_text(product.product_name),
        image.search_text,
    ).ratio()
    if name_similarity >= 0.52:
        similarity_score = round(name_similarity * 35, 2)
        score += similarity_score
        reasons.append(f"name_similarity:{name_similarity:.2f}")

    return round(score, 2), reasons


def confidence_from_scores(best_score: float, second_score: float) -> str:
    gap = best_score - second_score
    if best_score >= 95:
        return "high"
    if best_score >= 70 and gap >= 10:
        return "high"
    if best_score >= 52 and gap >= 8:
        return "medium"
    if best_score >= 36 and gap >= 10:
        return "low"
    return "unmatched"


def assign_images(images: Sequence[ImageAsset], products: Dict[str, ProductRecord]) -> List[ImageAssignment]:
    assignments: List[ImageAssignment] = []
    product_list = list(products.values())
    for image in images:
        candidates: List[MatchCandidate] = []
        for product in product_list:
            score, basis = score_image_for_product(image, product)
            if score <= 0:
                continue
            candidates.append(
                MatchCandidate(
                    product_id=product.product_id,
                    score=score,
                    confidence="",
                    basis=basis,
                )
            )

        candidates.sort(key=lambda item: item.score, reverse=True)
        top_candidates = candidates[:3]
        best = top_candidates[0] if top_candidates else None
        second_score = top_candidates[1].score if len(top_candidates) > 1 else 0.0
        if best is None:
            assignments.append(
                ImageAssignment(
                    image=image,
                    product_id=None,
                    score=0.0,
                    confidence="unmatched",
                    basis=[],
                    top_candidates=[],
                )
            )
            continue

        confidence = confidence_from_scores(best.score, second_score)
        if confidence == "unmatched":
            assignments.append(
                ImageAssignment(
                    image=image,
                    product_id=None,
                    score=best.score,
                    confidence="unmatched",
                    basis=[],
                    top_candidates=top_candidates,
                )
            )
            continue

        best.confidence = confidence
        assignments.append(
            ImageAssignment(
                image=image,
                product_id=best.product_id,
                score=best.score,
                confidence=confidence,
                basis=best.basis,
                top_candidates=top_candidates,
            )
        )
    return assignments


def bucket_product_matches(assignments: Sequence[ImageAssignment]) -> Dict[str, List[ImageAssignment]]:
    bucket: Dict[str, List[ImageAssignment]] = defaultdict(list)
    for assignment in assignments:
        if assignment.product_id:
            bucket[assignment.product_id].append(assignment)
    for rows in bucket.values():
        rows.sort(key=lambda item: item.score, reverse=True)
    return bucket


def list_to_json_cell(values: Sequence[str]) -> str:
    return json.dumps(list(values), ensure_ascii=False)


def path_to_asset_map(images: Sequence[ImageAsset]) -> Dict[str, ImageAsset]:
    return {image.path: image for image in images}


def classify_product_rows(
    products: Dict[str, ProductRecord],
    matches_by_product: Dict[str, List[ImageAssignment]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    product_rows: List[Dict[str, Any]] = []
    missing_rows: List[Dict[str, Any]] = []

    confidence_rank = {"low": 1, "medium": 2, "high": 3}
    inverse_confidence_rank = {value: key for key, value in confidence_rank.items()}

    for product_id in sorted(products, key=lambda key: (int(key) if key.isdigit() else 10**9, key)):
        product = products[product_id]
        matches = matches_by_product.get(product_id, [])
        by_type: Dict[str, List[ImageAssignment]] = defaultdict(list)
        for match in matches:
            by_type[match.image.detected_type].append(match)

        cover_candidates = by_type.get("cover", [])
        gallery_candidates = by_type.get("gallery", [])
        detail_candidates = by_type.get("detail", [])
        sku_candidates = by_type.get("sku", [])
        poster_candidates = by_type.get("poster", [])
        unknown_candidates = by_type.get("unknown", [])

        remote_cover_image = ""
        if product.existing_cover_images:
            remote_cover_image = product.existing_cover_images[0]
        elif product.existing_gallery_images:
            remote_cover_image = product.existing_gallery_images[0]
        elif product.sku_image_urls:
            remote_cover_image = product.sku_image_urls[0]

        remote_gallery_images = product.existing_gallery_images
        remote_detail_images = product.existing_detail_images
        remote_sku_images = product.sku_image_urls
        remote_poster_images: List[str] = []
        remote_unknown_images: List[str] = []

        local_cover_images = [item.image.path for item in cover_candidates]
        local_gallery_images = [item.image.path for item in gallery_candidates]
        local_detail_images = [item.image.path for item in detail_candidates]
        local_sku_images = [item.image.path for item in sku_candidates]
        local_poster_images = [item.image.path for item in poster_candidates]
        local_unknown_images = [item.image.path for item in unknown_candidates]

        basis_counter: Counter[str] = Counter()
        best_confidence = 0
        for match in matches:
            best_confidence = max(best_confidence, confidence_rank.get(match.confidence, 0))
            for basis in match.basis:
                basis_counter[basis] += 1

        match_confidence = inverse_confidence_rank.get(best_confidence, "low")
        match_basis = [basis for basis, _ in basis_counter.most_common(6)]

        notes: List[str] = []
        if not matches:
            notes.append("未匹配到本地素材图")
        if matches and all(match.confidence == "low" for match in matches):
            notes.append("仅存在低置信度匹配，请人工复核")
        if not remote_cover_image:
            notes.append("json 中缺少 cover 图路径")
        if not remote_detail_images:
            notes.append("json 中缺少 detail 图路径")

        row = {
            "product_id": product.product_id,
            "sku_id": list_to_json_cell(product.sku_ids),
            "product_name": product.product_name,
            "price": product.price,
            "original_price": product.original_price,
            "stock": product.stock,
            "category": product.category_name or product.category_id,
            "cover_image": remote_cover_image,
            "gallery_images": list_to_json_cell(remote_gallery_images),
            "detail_images": list_to_json_cell(remote_detail_images),
            "sku_images": list_to_json_cell(remote_sku_images),
            "poster_images": list_to_json_cell(remote_poster_images),
            "unknown_images": list_to_json_cell(remote_unknown_images),
            "matched_local_cover_files": list_to_json_cell(local_cover_images),
            "matched_local_gallery_files": list_to_json_cell(local_gallery_images),
            "matched_local_detail_files": list_to_json_cell(local_detail_images),
            "matched_local_sku_files": list_to_json_cell(local_sku_images),
            "matched_local_poster_files": list_to_json_cell(local_poster_images),
            "matched_local_unknown_files": list_to_json_cell(local_unknown_images),
            "matched_image_count": len(matches),
            "match_confidence": match_confidence,
            "match_basis": list_to_json_cell(match_basis),
            "source_json_file": product.source_json_file,
            "notes": "；".join(notes),
        }
        product_rows.append(row)

        if not matches or not remote_cover_image or not remote_detail_images:
            missing_rows.append(
                {
                    "product_id": product.product_id,
                    "product_name": product.product_name,
                    "matched_image_count": len(matches),
                    "has_json_cover": bool(remote_cover_image),
                    "has_json_gallery": bool(remote_gallery_images),
                    "has_json_detail": bool(remote_detail_images),
                    "has_json_sku": bool(remote_sku_images),
                    "has_local_cover": bool(local_cover_images),
                    "has_local_gallery": bool(local_gallery_images),
                    "has_local_detail": bool(local_detail_images),
                    "has_local_sku": bool(local_sku_images),
                    "source_json_file": product.source_json_file,
                    "notes": "；".join(notes),
                }
            )

    return product_rows, missing_rows


def build_detail_image_rows(
    products: Dict[str, ProductRecord],
    matches_by_product: Dict[str, List[ImageAssignment]],
    image_asset_map: Dict[str, ImageAsset],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for product_id in sorted(products, key=lambda key: (int(key) if key.isdigit() else 10**9, key)):
        product = products[product_id]
        matches = matches_by_product.get(product_id, [])
        detail_matches = [match for match in matches if match.image.detected_type == "detail"]

        linked_local_indices = set()
        remote_detail_urls = product.existing_detail_images or []
        for index, url in enumerate(remote_detail_urls, start=1):
            matched_local = None
            remote_stem = normalize_text(extract_url_stem(url))
            for local_index, match in enumerate(detail_matches):
                if local_index in linked_local_indices:
                    continue
                local_stem = normalize_text(match.image.stem)
                if remote_stem and remote_stem == local_stem:
                    matched_local = (local_index, match)
                    break
            if matched_local is None and len(remote_detail_urls) == len(detail_matches) and detail_matches:
                candidate = detail_matches[index - 1]
                matched_local = (index - 1, candidate)
            if matched_local is not None:
                local_index, match = matched_local
                linked_local_indices.add(local_index)
                asset = match.image
                rows.append(
                    {
                        "product_id": product.product_id,
                        "product_name": product.product_name,
                        "category": product.category_name or product.category_id,
                        "detail_index": index,
                        "detail_preview": "",
                        "json_image_path": url,
                        "matched_local_path": asset.path,
                        "relative_path": asset.relative_path,
                        "extension": asset.ext,
                        "size_bytes": asset.size_bytes,
                        "width": asset.width,
                        "height": asset.height,
                        "detected_type": asset.detected_type,
                        "match_confidence": match.confidence,
                        "match_basis": list_to_json_cell(match.basis),
                        "source_type": "json_detail_with_local_match",
                        "source_json_file": product.source_json_file,
                        "notes": "",
                    }
                )
            else:
                rows.append(
                    {
                        "product_id": product.product_id,
                        "product_name": product.product_name,
                        "category": product.category_name or product.category_id,
                        "detail_index": index,
                        "detail_preview": "",
                        "json_image_path": url,
                        "matched_local_path": "",
                        "relative_path": "",
                        "extension": Path(extract_url_basename(url)).suffix.lower(),
                        "size_bytes": "",
                        "width": "",
                        "height": "",
                        "detected_type": "detail",
                        "match_confidence": "unmatched",
                        "match_basis": list_to_json_cell(["existing_detail_images"]),
                        "source_type": "json_detail_only",
                        "source_json_file": product.source_json_file,
                        "notes": "json 中存在详情图路径，但未匹配到本地素材",
                    }
                )

        remaining_local_details = [
            (idx, match)
            for idx, match in enumerate(detail_matches)
            if idx not in linked_local_indices
        ]
        for extra_index, (_, match) in enumerate(remaining_local_details, start=1):
            asset = match.image
            rows.append(
                {
                    "product_id": product.product_id,
                    "product_name": product.product_name,
                    "category": product.category_name or product.category_id,
                    "detail_index": f"local-{extra_index}",
                    "detail_preview": "",
                    "json_image_path": "",
                    "matched_local_path": asset.path,
                    "relative_path": asset.relative_path,
                    "extension": asset.ext,
                    "size_bytes": asset.size_bytes,
                    "width": asset.width,
                    "height": asset.height,
                    "detected_type": asset.detected_type,
                    "match_confidence": match.confidence,
                    "match_basis": list_to_json_cell(match.basis),
                    "source_type": "local_detail_without_json_path",
                    "source_json_file": product.source_json_file,
                    "notes": "本地匹配到详情图，但 json 中没有对应详情图路径",
                }
            )

        if not remote_detail_urls and not detail_matches:
            rows.append(
                {
                    "product_id": product.product_id,
                    "product_name": product.product_name,
                    "category": product.category_name or product.category_id,
                    "detail_index": "",
                    "detail_preview": "",
                    "json_image_path": "",
                    "matched_local_path": "",
                    "relative_path": "",
                    "extension": "",
                    "size_bytes": "",
                    "width": "",
                    "height": "",
                    "detected_type": "detail",
                    "match_confidence": "missing",
                    "match_basis": list_to_json_cell([]),
                    "source_type": "missing_detail",
                    "source_json_file": product.source_json_file,
                    "notes": "缺少详情图",
                }
            )

    return rows


def build_product_sheet_rows(
    product_rows: Sequence[Dict[str, Any]],
    matches_by_product: Dict[str, List[ImageAssignment]],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row in product_rows:
        product_id = row["product_id"]
        matches = matches_by_product.get(product_id, [])
        detail_matches = [match for match in matches if match.image.detected_type == "detail"]
        new_row = dict(row)
        new_row["cover_preview"] = ""
        new_row["detail_preview"] = ""
        new_row["detail_image_count"] = len(detail_matches)
        rows.append(new_row)
    return rows


def build_image_embeds_for_products(
    product_rows: Sequence[Dict[str, Any]],
    matches_by_product: Dict[str, List[ImageAssignment]],
) -> List[ImageEmbed]:
    embeds: List[ImageEmbed] = []
    for row_number, row in enumerate(product_rows, start=2):
        product_id = row["product_id"]
        local_cover_candidates = collect_string_list(row.get("matched_local_cover_files"))
        local_gallery_candidates = collect_string_list(row.get("matched_local_gallery_files"))
        local_cover_path = ""
        if local_cover_candidates:
            local_cover_path = local_cover_candidates[0]
        elif local_gallery_candidates:
            local_cover_path = local_gallery_candidates[0]
        if local_cover_path and Path(local_cover_path).exists():
            embeds.append(
                ImageEmbed(
                    row_number=row_number,
                    column_name="cover_preview",
                    image_path=local_cover_path,
                )
            )
        matches = matches_by_product.get(product_id, [])
        detail_match = next((match for match in matches if match.image.detected_type == "detail"), None)
        if detail_match and Path(detail_match.image.path).exists():
            embeds.append(
                ImageEmbed(
                    row_number=row_number,
                    column_name="detail_preview",
                    image_path=detail_match.image.path,
                )
            )
    return embeds


def build_image_embeds_for_detail_rows(detail_rows: Sequence[Dict[str, Any]]) -> List[ImageEmbed]:
    embeds: List[ImageEmbed] = []
    for row_number, row in enumerate(detail_rows, start=2):
        image_path = row.get("matched_local_path") or ""
        if image_path and Path(image_path).exists():
            embeds.append(
                ImageEmbed(
                    row_number=row_number,
                    column_name="detail_preview",
                    image_path=image_path,
                )
            )
    return embeds


def build_unmatched_rows(assignments: Sequence[ImageAssignment]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for assignment in assignments:
        if assignment.product_id is not None:
            continue
        top_candidates = [
            {
                "product_id": candidate.product_id,
                "score": candidate.score,
                "basis": candidate.basis,
            }
            for candidate in assignment.top_candidates
        ]
        rows.append(
            {
                "file_name": assignment.image.file_name,
                "path": assignment.image.path,
                "extension": assignment.image.ext,
                "size_bytes": assignment.image.size_bytes,
                "sha1": assignment.image.sha1,
                "detected_type": assignment.image.detected_type,
                "top_candidates": json.dumps(top_candidates, ensure_ascii=False),
            }
        )
    return rows


def build_duplicate_rows(images: Sequence[ImageAsset]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[ImageAsset]] = defaultdict(list)
    for image in images:
        grouped[image.sha1].append(image)

    rows: List[Dict[str, Any]] = []
    group_index = 1
    for sha1_value, group in sorted(grouped.items()):
        if len(group) <= 1:
            continue
        for image in group:
            rows.append(
                {
                    "duplicate_group": group_index,
                    "sha1": sha1_value,
                    "size_bytes": image.size_bytes,
                    "file_name": image.file_name,
                    "path": image.path,
                    "detected_type": image.detected_type,
                    "duplicate_count": len(group),
                }
            )
        group_index += 1
    return rows


def write_csv(path: Path, rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def autosize_sheet(worksheet) -> None:
    widths: Dict[int, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 12), 60)


def embed_images(
    worksheet,
    fieldnames: Sequence[str],
    embeds: Sequence[ImageEmbed],
    preview_size: Tuple[int, int] = (110, 110),
) -> None:
    if not embeds:
        return
    column_map = {name: idx + 1 for idx, name in enumerate(fieldnames)}
    row_heights: Dict[int, float] = {}
    for embed in embeds:
        column_index = column_map.get(embed.column_name)
        if column_index is None:
            continue
        image_path = Path(embed.image_path)
        if not image_path.exists():
            continue
        try:
            image = XLImage(str(image_path))
            max_width, max_height = preview_size
            width = getattr(image, "width", None) or max_width
            height = getattr(image, "height", None) or max_height
            scale = min(max_width / width, max_height / height, 1)
            image.width = max(1, int(width * scale))
            image.height = max(1, int(height * scale))
            cell = f"{get_column_letter(column_index)}{embed.row_number}"
            image.anchor = cell
            worksheet.add_image(image)
            current_height = row_heights.get(embed.row_number, worksheet.row_dimensions[embed.row_number].height or 15)
            row_heights[embed.row_number] = max(current_height, image.height * 0.78)
            worksheet.column_dimensions[get_column_letter(column_index)].width = max(
                worksheet.column_dimensions[get_column_letter(column_index)].width or 10,
                18,
            )
        except Exception:
            logging.exception("Failed to embed image into workbook: %s", image_path)
    for row_number, height in row_heights.items():
        worksheet.row_dimensions[row_number].height = min(max(height, 80), 140)


def write_xlsx(
    path: Path,
    product_rows: Sequence[Dict[str, Any]],
    detail_rows: Sequence[Dict[str, Any]],
    unmatched_rows: Sequence[Dict[str, Any]],
    duplicate_rows: Sequence[Dict[str, Any]],
    missing_rows: Sequence[Dict[str, Any]],
    audit_summary_rows: Sequence[Dict[str, Any]],
    sheet_embeds: Optional[Dict[str, List[ImageEmbed]]] = None,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = [
        ("products", product_rows),
        ("detail_images", detail_rows),
        ("unmatched_images", unmatched_rows),
        ("duplicate_images", duplicate_rows),
        ("missing_images", missing_rows),
        ("audit_summary", audit_summary_rows),
    ]

    for sheet_name, rows in sheets:
        worksheet = workbook.create_sheet(title=sheet_name)
        fieldnames: List[str] = []
        for row in rows:
            for key in row.keys():
                if key not in fieldnames:
                    fieldnames.append(key)
        if not fieldnames:
            fieldnames = ["message"]
            rows = [{"message": "no data"}]

        worksheet.append(fieldnames)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        for row in rows:
            worksheet.append([row.get(column, "") for column in fieldnames])

        for data_row in worksheet.iter_rows(min_row=2):
            for cell in data_row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.freeze_panes = "A2"
        autosize_sheet(worksheet)
        if sheet_embeds:
            embed_images(worksheet, fieldnames, sheet_embeds.get(sheet_name, []))

    workbook.save(path)


def build_audit_summary_rows(
    args: argparse.Namespace,
    all_data_files: Sequence[Path],
    datasets: Dict[str, List[Path]],
    images: Sequence[ImageAsset],
    product_rows: Sequence[Dict[str, Any]],
    detail_rows: Sequence[Dict[str, Any]],
    unmatched_rows: Sequence[Dict[str, Any]],
    duplicate_rows: Sequence[Dict[str, Any]],
    missing_rows: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    summary_rows: List[Dict[str, Any]] = [
        {"section": "summary", "key": "generated_at", "value": utc_now()},
        {"section": "summary", "key": "image_roots", "value": list_to_json_cell([str(path) for path in args.image_root])},
        {"section": "summary", "key": "data_roots", "value": list_to_json_cell([str(path) for path in args.data_root])},
        {"section": "summary", "key": "output_dir", "value": str(args.output_dir)},
        {"section": "summary", "key": "json_jsonl_files_scanned", "value": len(all_data_files)},
        {"section": "summary", "key": "image_files_scanned", "value": len(images)},
        {"section": "summary", "key": "products_detected", "value": len(product_rows)},
        {"section": "summary", "key": "detail_image_rows", "value": len(detail_rows)},
        {"section": "summary", "key": "unmatched_images", "value": len(unmatched_rows)},
        {"section": "summary", "key": "duplicate_rows", "value": len(duplicate_rows)},
        {"section": "summary", "key": "missing_image_products", "value": len(missing_rows)},
    ]

    for dataset_name, files in sorted(datasets.items()):
        for file_path in sorted(files):
            summary_rows.append(
                {
                    "section": f"dataset:{dataset_name}",
                    "key": "source_file",
                    "value": str(file_path.resolve()),
                }
            )
    return summary_rows


def build_mapping_audit(
    args: argparse.Namespace,
    all_data_files: Sequence[Path],
    datasets: Dict[str, List[Path]],
    images: Sequence[ImageAsset],
    products: Dict[str, ProductRecord],
    assignments: Sequence[ImageAssignment],
    product_rows: Sequence[Dict[str, Any]],
    detail_rows: Sequence[Dict[str, Any]],
    unmatched_rows: Sequence[Dict[str, Any]],
    duplicate_rows: Sequence[Dict[str, Any]],
    missing_rows: Sequence[Dict[str, Any]],
) -> Dict[str, Any]:
    matches_by_product = bucket_product_matches(assignments)
    return {
        "generated_at": utc_now(),
        "image_roots": [str(path) for path in args.image_root],
        "data_roots": [str(path) for path in args.data_root],
        "output_dir": str(args.output_dir),
        "summary": {
            "json_jsonl_files_scanned": len(all_data_files),
            "image_files_scanned": len(images),
            "products_detected": len(products),
            "matched_products": sum(1 for row in product_rows if row["matched_image_count"] > 0),
            "detail_image_rows": len(detail_rows),
            "unmatched_images": len(unmatched_rows),
            "duplicate_rows": len(duplicate_rows),
            "missing_image_products": len(missing_rows),
        },
        "dataset_files": {
            key: [str(path.resolve()) for path in sorted(value)]
            for key, value in datasets.items()
        },
        "products": [
            {
                "product_id": product.product_id,
                "product_name": product.product_name,
                "category": product.category_name or product.category_id,
                "source_json_file": product.source_json_file,
                "merged_sources": product.merged_sources,
                "aliases": product.aliases[:20],
                "unique_aliases": product.unique_aliases[:20],
                "matched_images": [
                    {
                        "path": match.image.path,
                        "detected_type": match.image.detected_type,
                        "score": match.score,
                        "confidence": match.confidence,
                        "basis": match.basis,
                    }
                    for match in matches_by_product.get(product.product_id, [])
                ],
            }
            for product in products.values()
        ],
        "detail_images": detail_rows,
        "unmatched_images": unmatched_rows,
        "duplicate_images": duplicate_rows,
        "missing_images_products": missing_rows,
    }


def determine_default_image_roots(repo_root: Path) -> List[Path]:
    family_root = repo_root.parents[1]
    candidates = [
        family_root / "prod/问兰素材/小程序组合装",
        family_root / "prod/问兰素材/素材2-24",
        family_root / "prod/问兰素材/蔚总照片素材",
    ]
    return [path for path in candidates if path.exists()]


def parse_args() -> argparse.Namespace:
    repo_root = Path(__file__).resolve().parents[1]
    default_image_roots = determine_default_image_roots(repo_root)
    parser = argparse.ArgumentParser(
        description="整理商品素材图，并与项目 json/jsonl 商品数据对齐导出为 Excel/CSV。"
    )
    parser.add_argument(
        "--image-root",
        action="append",
        type=Path,
        default=default_image_roots,
        help="素材图片根目录，可重复传入；默认使用已知素材目录。",
    )
    parser.add_argument(
        "--data-root",
        action="append",
        type=Path,
        default=[repo_root],
        help="递归扫描 json/jsonl 的根目录，可重复传入；默认扫描当前仓库。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=repo_root / "artifacts/product_image_mapping",
        help="输出目录。",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="输出更详细日志。",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.output_dir, args.verbose)

    logging.info("Output directory: %s", args.output_dir)
    logging.info("Image roots: %s", [str(path) for path in args.image_root])
    logging.info("Data roots: %s", [str(path) for path in args.data_root])

    images = scan_images(args.image_root)
    all_data_files, datasets = scan_data_files(args.data_root, args.output_dir)

    category_map = merge_category_maps(datasets.get("categories", []))
    products = merge_products(datasets.get("products", []), category_map)
    skus = merge_skus(datasets.get("skus", []))
    attach_skus_to_products(products, skus)
    build_aliases(products)

    logging.info("Normalized products: %s", len(products))
    logging.info("Normalized skus: %s", len(skus))

    assignments = assign_images(images, products)
    matches_by_product = bucket_product_matches(assignments)
    product_rows, missing_rows = classify_product_rows(products, matches_by_product)
    detail_rows = build_detail_image_rows(products, matches_by_product, path_to_asset_map(images))
    product_sheet_rows = build_product_sheet_rows(product_rows, matches_by_product)
    unmatched_rows = build_unmatched_rows(assignments)
    duplicate_rows = build_duplicate_rows(images)
    audit_summary_rows = build_audit_summary_rows(
        args,
        all_data_files,
        datasets,
        images,
        product_rows,
        detail_rows,
        unmatched_rows,
        duplicate_rows,
        missing_rows,
    )
    audit_payload = build_mapping_audit(
        args,
        all_data_files,
        datasets,
        images,
        products,
        assignments,
        product_rows,
        detail_rows,
        unmatched_rows,
        duplicate_rows,
        missing_rows,
    )

    args.output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(args.output_dir / "products.csv", product_rows)
    write_csv(args.output_dir / "unmatched_images.csv", unmatched_rows)
    write_csv(args.output_dir / "duplicate_images.csv", duplicate_rows)
    write_csv(args.output_dir / "missing_images_products.csv", missing_rows)
    write_xlsx(
        args.output_dir / "product_image_mapping.xlsx",
        product_sheet_rows,
        detail_rows,
        unmatched_rows,
        duplicate_rows,
        missing_rows,
        audit_summary_rows,
        sheet_embeds={
            "products": build_image_embeds_for_products(product_sheet_rows, matches_by_product),
            "detail_images": build_image_embeds_for_detail_rows(detail_rows),
        },
    )
    with (args.output_dir / "mapping_audit.json").open("w", encoding="utf-8") as handle:
        json.dump(audit_payload, handle, ensure_ascii=False, indent=2)

    logging.info("Generated xlsx/csv/json outputs under %s", args.output_dir)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception as exc:  # pragma: no cover - top-level guard
        traceback.print_exc()
        logging.error("Fatal error: %s", exc)
        raise SystemExit(1)
